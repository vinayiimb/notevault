import json
import openpyxl

XLSX_PATH = '/Users/sayam/Downloads/shivaji_papers_syllabus_mapping.xlsx'
JSON_PATH = 'src/data/du-question-bank-full-mapped.json'

print("Loading existing question bank data...")
with open(JSON_PATH, 'r', encoding='utf-8') as f:
    qb_data = json.load(f)

initial_count = len(qb_data)
print(f"Loaded {initial_count} existing papers.")

print("Loading Shivaji mapped papers from Excel...")
wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws = wb['All Papers Mapped']
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

# Mappings fields to the standard RawQuestionBankRow schema
# Keys needed: ['officialProgramme', 'semester', 'paperType', 'subjectPaperName', 'courseNumber', 'upc', 'credits', 'matchedCategories', 'sourceType', 'officialPageUrl', 'officialPaperLink', 'questionPaperLink', 'questionPaperSession', 'questionPaperYear', 'questionPaperSet', 'questionPaperMarks', 'matchSource', 'recoveredUpc', 'isShivaji']

shivaji_added = 0
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
    d = dict(zip(headers, row))
    
    prog = d.get('Matched Programme(s)')
    subj = d.get('Matched Official Paper Name')
    link = d.get('Direct Download Link') or d.get('Drive View Link')
    
    # We only import the mapped papers (2,328 entries)
    if not (prog and subj and link):
        continue
    
    # Clean the programme field: if it's a comma-separated list of multiple programmes, we handle them or use it directly
    # Typically they are semicolon/comma separated or simple text. Let's split and add for each programme to map correctly!
    programmes_to_add = [p.strip() for p in str(prog).split(';') if p.strip()]
    if not programmes_to_add:
        programmes_to_add = [str(prog).strip()]
        
    for p_name in programmes_to_add:
        row_obj = {
            "officialProgramme": p_name,
            "semester": str(d.get('Semester') or "").strip() or None,
            "paperType": str(d.get('Paper Type') or "").strip() or None,
            "subjectPaperName": str(subj).strip(),
            "courseNumber": str(d.get('Matched Course Number') or "").strip() or None,
            "upc": str(d.get('Matched UPC') or "").strip() or None,
            "credits": str(d.get('Matched Credits') or "").strip() or None,
            "matchedCategories": None,
            "sourceType": str(d.get('Paper Type') or "").strip() or None,
            "officialPageUrl": None,
            "officialPaperLink": None,
            "questionPaperLink": str(link).strip(),
            "questionPaperSession": str(d.get('Exam Session') or "").strip() or None,
            "questionPaperYear": str(d.get('Year') or "").strip() or None,
            "questionPaperSet": None,
            "questionPaperMarks": None,
            "matchSource": "Shivaji",
            "recoveredUpc": None,
            "isShivaji": True
        }
        qb_data.append(row_obj)
        shivaji_added += 1

print(f"Added {shivaji_added} Shivaji paper-programme mappings.")
print(f"New total paper count: {len(qb_data)}.")

print("Saving updated question bank database...")
with open(JSON_PATH, 'w', encoding='utf-8') as f:
    json.dump(qb_data, f, ensure_ascii=False, indent=2)

print("Saved successfully!")
