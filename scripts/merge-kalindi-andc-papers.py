import json
import openpyxl

XLSX_PATH = '/Users/sayam/Downloads/kalindi_andc_master_papers_final11.xlsx'
JSON_PATH = 'src/data/du-question-bank-full-mapped.json'

print("Loading existing question bank data...")
with open(JSON_PATH, 'r', encoding='utf-8') as f:
    qb_data = json.load(f)

initial_count = len(qb_data)
print(f"Loaded {initial_count} existing papers.")

print("Loading Kalindi & ANDC master papers from Excel...")
wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws = wb['All Papers']
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

kalindi_count = 0
andc_count = 0
added_count = 0

for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
    d = dict(zip(headers, row))
    
    prog = d.get('Course')
    subj = d.get('Subject Name')
    link = d.get('Drive Link')
    college_raw = str(d.get('College') or "").strip()
    
    if not (prog and subj and link):
        continue
    
    # Determine college marker
    if "Kalindi" in college_raw:
        college = "Kalindi"
        is_kalindi = True
        is_andc = False
        kalindi_count += 1
    elif "Acharya" in college_raw or "ANDC" in college_raw:
        college = "ANDC"
        is_kalindi = False
        is_andc = True
        andc_count += 1
    else:
        college = "Other"
        is_kalindi = False
        is_andc = False
    
    # Clean programmes: if semicolon separated, handle each
    programmes_to_add = [p.strip() for p in str(prog).split(';') if p.strip()]
    if not programmes_to_add:
        programmes_to_add = [str(prog).strip()]
        
    for p_name in programmes_to_add:
        row_obj = {
            "officialProgramme": p_name,
            "semester": str(d.get('Semester') or "").strip() or None,
            "paperType": str(d.get('Paper Type') or "").strip() or None,
            "subjectPaperName": str(subj).strip(),
            "courseNumber": str(d.get('Course Number') or "").strip() or None,
            "upc": str(d.get('Code') or "").strip() or None,
            "credits": str(d.get('Credits') or "").strip() or None,
            "matchedCategories": None,
            "sourceType": str(d.get('Paper Type') or "").strip() or None,
            "officialPageUrl": None,
            "officialPaperLink": str(d.get('Official Paper Link') or "").strip() or None,
            "questionPaperLink": str(link).strip(),
            "questionPaperSession": None,
            "questionPaperYear": str(d.get('Year') or "").strip() or None,
            "questionPaperSet": None,
            "questionPaperMarks": None,
            "matchSource": college_raw,
            "recoveredUpc": None,
            "college": college,
            "isKalindi": is_kalindi,
            "isANDC": is_andc,
            "isShivaji": False
        }
        qb_data.append(row_obj)
        added_count += 1

print(f"Added {kalindi_count} Kalindi papers and {andc_count} ANDC papers ({added_count} total new mapped entries).")
print(f"New total paper count in master file: {len(qb_data)}.")

print("Writing to du-question-bank-full-mapped.json...")
with open(JSON_PATH, 'w', encoding='utf-8') as f:
    json.dump(qb_data, f, ensure_ascii=False, indent=2)

print("Saved successfully!")
